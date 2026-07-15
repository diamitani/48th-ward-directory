#!/usr/bin/env python3
"""Data pipeline: Extract + clean 48th Ward spreadsheets into structured JSON/CSV."""
import openpyxl, json, csv, re, os
from pathlib import Path

OUT = Path("/Users/patmini/Civic/48th-ward-directory/data")

def clean_phone(p):
    """Normalize phone numbers."""
    if not p: return None
    p = str(p).strip()
    p = re.sub(r'[^\d+\-\(\) ]', '', p)
    if len(p) >= 10:
        digits = re.sub(r'\D', '', p)
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        if len(digits) == 11 and digits[0] == '1':
            return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return p

def clean_website(w):
    if not w: return None
    w = str(w).strip()
    if w and not w.startswith('http'):
        w = 'https://' + w
    return w

def dedupe_rows(rows, key_fn):
    """Deduplicate rows by key, keeping the one with most non-None fields."""
    seen = {}
    for r in rows:
        k = key_fn(r)
        if k is None: continue
        if k not in seen:
            seen[k] = r
        else:
            # Keep the row with more non-None fields
            if sum(1 for v in r.values() if v) > sum(1 for v in seen[k].values() if v):
                seen[k] = r
    return list(seen.values())

# ─── Load Master Spreadsheet ───
MS = "/Users/patmini/Downloads/48th_Ward_Master_Spreadsheet.xlsx"
wb = openpyxl.load_workbook(MS, data_only=True)

# ─── 1. BUSINESSES ───
print("Processing businesses...")
ws = wb["Business Directory"]
headers = [c.value for c in ws[1]]
businesses = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if not d.get("Name"): continue
    b = {
        "precinct": str(d.get("Precinct", "")).strip(),
        "category": str(d.get("Category", "")).strip() if d.get("Category") else None,
        "name": str(d.get("Name", "")).strip(),
        "address": str(d.get("Address", "")).strip() if d.get("Address") else None,
        "neighborhood": str(d.get("Neighborhood", "")).strip() if d.get("Neighborhood") else None,
        "phone": clean_phone(d.get("Phone")),
        "website": clean_website(d.get("Website")),
        "email": str(d.get("Email", "")).strip() if d.get("Email") else None,
        "hours": str(d.get("Hours", "")).strip() if d.get("Hours") else None,
        "owner": str(d.get("Owner", "")).strip() if d.get("Owner") else None,
        "owner_phone": clean_phone(d.get("Owner Phone")),
        "community_friendly": str(d.get("Community Friendly", "")).strip() == "Yes",
        "notes": str(d.get("Notes", "")).strip() if d.get("Notes") else None,
    }
    businesses.append(b)

# Deduplicate by name+address
businesses = dedupe_rows(businesses, lambda r: (r["name"].lower(), r.get("address","")))
print(f"  → {len(businesses)} unique businesses after dedup")

# ─── 2. PRECINCTS ───
print("Processing precincts...")
ws = wb["Precinct Directory"]
precincts = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    p = str(row[0]).strip()
    if not p.startswith("P"): continue
    precincts.append({
        "precinct": p,
        "neighborhood": str(row[1]).strip() if row[1] else None,
        "businesses_count": int(row[5]) if row[5] else 0,
        "schools_count": int(row[6]) if row[6] else 0,
        "religious_count": int(row[7]) if row[7] else 0,
        "parks_count": int(row[8]) if row[8] else 0,
        "gov_buildings_count": int(row[9]) if row[9] else 0,
        "services_count": int(row[10]) if row[10] else 0,
        "residential_buildings": int(row[11]) if row[11] else 0,
    })
print(f"  → {len(precincts)} precincts")

# ─── 3. PRECINCT CAPTAINS ───
print("Processing precinct captains...")
PC = "/Users/patmini/Downloads/2026 Precinct Captains 48th Ward.xlsx"
wb2 = openpyxl.load_workbook(PC, data_only=True)
ws2 = wb2["PCs"]
headers2 = [c.value for c in ws2[1]]
captains = []
current_precinct = None
current_location = None
for row in ws2.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers2, row))
    if d.get("Polling Location"):
        current_location = str(d["Polling Location"]).strip()
    if d.get("Precinct") is not None:
        try: current_precinct = int(d["Precinct"])
        except: current_precinct = None
    name = d.get("Precinct Captain")
    if not name: continue
    captains.append({
        "precinct": current_precinct,
        "polling_location": current_location,
        "name": str(name).strip(),
        "email": str(d.get("Email", "")).strip() if d.get("Email") else None,
        "phone": clean_phone(d.get("Phone")),
        "google_group": str(d.get("google group?", "")).strip() == "y" if d.get("google group?") else False,
        "signal": str(d.get("Signal?", "")).strip() == "y" if d.get("Signal?") else False,
        "vote_builder": str(d.get("Vote Builder?", "")).strip() == "y" if d.get("Vote Builder?") else False,
        "notes": str(d.get("Notes", "")).strip() if d.get("Notes") else None,
    })
print(f"  → {len(captains)} captains")

# Group captains by precinct
from collections import defaultdict
precinct_captains = defaultdict(list)
for c in captains:
    precinct_captains[c["precinct"]].append(c)

# ─── 4. TURNOUT DATA ───
print("Processing turnout data...")
ws3 = wb2["Turnout 2026 Primary"]
turnout = []
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    try: p = int(row[1])
    except: continue
    turnout.append({
        "precinct": p,
        "registered_voters": int(row[2]) if row[2] else None,
        "ballots_cast": int(row[3]) if row[3] else None,
        "turnout_pct": None if not row[4] else (float(str(row[4]).replace('%','')) if row[4] else None),
    })
print(f"  → {len(turnout)} precinct turnout entries")

# ─── 5. SCHOOLS ───
print("Processing schools...")
ws = wb["Schools Directory"]
headers = [c.value for c in ws[1]]
schools = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if not d.get("School Name"): continue
    schools.append({
        "precinct": str(d.get("Precinct", "")).strip(),
        "name": str(d["School Name"]).strip(),
        "type": str(d.get("Type", "")).strip() if d.get("Type") else None,
        "level": str(d.get("Level", "")).strip() if d.get("Level") else None,
        "address": str(d.get("Address", "")).strip() if d.get("Address") else None,
        "phone": clean_phone(d.get("Phone")),
        "website": clean_website(d.get("Website")),
        "principal": str(d.get("Principal", "")).strip() if d.get("Principal") else None,
        "principal_email": str(d.get("Principal Email", "")).strip() if d.get("Principal Email") else None,
        "enrollment": int(d["Enrollment"]) if d.get("Enrollment") else None,
    })
schools = dedupe_rows(schools, lambda r: r["name"].lower())
print(f"  → {len(schools)} unique schools")

# ─── 6. RELIGIOUS INSTITUTIONS ───
print("Processing religious institutions...")
ws = wb["Religious Institutions"]
headers = [c.value for c in ws[1]]
religious = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if not d.get("Name"): continue
    religious.append({
        "precinct": str(d.get("Precinct", "")).strip(),
        "name": str(d["Name"]).strip(),
        "religion": str(d.get("Religion", "")).strip() if d.get("Religion") else None,
        "denomination": str(d.get("Denomination", "")).strip() if d.get("Denomination") else None,
        "address": str(d.get("Address", "")).strip() if d.get("Address") else None,
        "phone": clean_phone(d.get("Phone")),
        "website": clean_website(d.get("Website")),
        "leader": str(d.get("Leader", "")).strip() if d.get("Leader") else None,
        "leader_title": str(d.get("Leader Title", "")).strip() if d.get("Leader Title") else None,
        "congregation_size": str(d.get("Congregation Size", "")).strip() if d.get("Congregation Size") else None,
        "community_programs": str(d.get("Community Programs", "")).strip() if d.get("Community Programs") else None,
    })
religious = dedupe_rows(religious, lambda r: r["name"].lower())
print(f"  → {len(religious)} unique religious institutions")

# ─── 7. PARKS ───
ws = wb["Parks & Public Spaces"]
headers = [c.value for c in ws[1]]
parks = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if not d.get("Name"): continue
    parks.append({
        "precinct": str(d.get("Precinct", "")).strip(),
        "name": str(d["Name"]).strip(),
        "type": str(d.get("Type", "")).strip() if d.get("Type") else None,
        "address": str(d.get("Address", "")).strip() if d.get("Address") else None,
    })
parks = dedupe_rows(parks, lambda r: r["name"].lower())
print(f"  → {len(parks)} unique parks")

# ─── 8. COMMUNITY SERVICES ───
ws = wb["Community Services"]
headers = [c.value for c in ws[1]]
services = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if not d.get("Name"): continue
    services.append({
        "precinct": str(d.get("Precinct", "")).strip(),
        "name": str(d["Name"]).strip(),
        "type": str(d.get("Type", "")).strip() if d.get("Type") else None,
        "address": str(d.get("Address", "")).strip() if d.get("Address") else None,
        "phone": clean_phone(d.get("Phone")),
        "website": clean_website(d.get("Website")),
    })
services = dedupe_rows(services, lambda r: r["name"].lower())
print(f"  → {len(services)} unique community services")

# ─── WRITE JSON ───
Path(OUT / "json").mkdir(parents=True, exist_ok=True)
Path(OUT / "csv").mkdir(parents=True, exist_ok=True)

datasets = {
    "businesses": businesses,
    "precincts": precincts,
    "captains": captains,
    "turnout": turnout,
    "schools": schools,
    "religious": religious,
    "parks": parks,
    "services": services,
}

for name, data in datasets.items():
    # JSON
    with open(OUT / "json" / f"{name}.json", "w") as f:
        json.dump(data, f, indent=2, default=str)
    
    # CSV
    if data:
        with open(OUT / "csv" / f"{name}.csv", "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

# ─── SUMMARY ───
print("\n=== DATA PIPELINE COMPLETE ===")
print(f"Businesses:       {len(businesses)}")
print(f"Precincts:        {len(precincts)}")
print(f"Captains:         {len(captains)}")
print(f"Turnout entries:  {len(turnout)}")
print(f"Schools:          {len(schools)}")
print(f"Religious:        {len(religious)}")
print(f"Parks:            {len(parks)}")
print(f"Services:         {len(services)}")
print(f"\nCategory breakdown:")
from collections import Counter
cats = Counter(b["category"] for b in businesses if b["category"])
for cat, count in cats.most_common(20):
    print(f"  {cat}: {count}")
print(f"\nFiles written to {OUT}/json/ and {OUT}/csv/")
