import openpyxl, json, sys

# File 1: Master Spreadsheet
print("=== 48TH WARD MASTER SPREADSHEET ===")
wb = openpyxl.load_workbook('/Users/patmini/Downloads/48th_Ward_Master_Spreadsheet.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(f'\n--- Sheet: {s} (rows: {ws.max_row}, cols: {ws.max_column}) ---')
    headers = [c.value for c in ws[1]]
    print('Headers:', headers)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(12, ws.max_row), values_only=True)):
        print(f'Row {i+2}:', list(row))

print("\n\n=== 2026 PRECINCT CAPTAINS ===")
wb2 = openpyxl.load_workbook('/Users/patmini/Downloads/2026 Precinct Captains 48th Ward.xlsx', data_only=True)
print('Sheets:', wb2.sheetnames)
for s in wb2.sheetnames:
    ws = wb2[s]
    print(f'\n--- Sheet: {s} (rows: {ws.max_row}, cols: {ws.max_column}) ---')
    headers = [c.value for c in ws[1]]
    print('Headers:', headers)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(45, ws.max_row), values_only=True)):
        print(f'Row {i+2}:', list(row))
